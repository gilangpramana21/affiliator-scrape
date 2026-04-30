#!/usr/bin/env python3
"""
Web Dashboard for Tokopedia Affiliator Scraper
Flask-based dashboard with scraping controls
"""
from flask import Flask, render_template, request, jsonify, send_file
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
import subprocess
import threading
import re

app = Flask(__name__)

# Global state
scraping_status = {
    'is_running': False,
    'progress': 0,
    'total': 0,
    'current_creator': '',
    'message': 'Ready'
}


def load_data():
    """Load scraped data"""
    json_path = Path("output/affiliators_full.json")
    
    if not json_path.exists():
        return pd.DataFrame()
    
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    return pd.DataFrame(data)


def remove_duplicates(df):
    """Remove duplicate creators based on username or creator ID"""
    if df.empty:
        return df
    
    # Remove duplicates based on username (keep first occurrence)
    if 'username' in df.columns:
        df = df.drop_duplicates(subset=['username'], keep='first')
    
    # Also check for duplicate emails/whatsapp
    if 'email' in df.columns:
        # Remove rows with duplicate emails (but keep rows with null emails)
        df = df[~(df['email'].notna() & df['email'].duplicated(keep='first'))]
    
    if 'whatsapp' in df.columns:
        # Remove rows with duplicate whatsapp (but keep rows with null whatsapp)
        df = df[~(df['whatsapp'].notna() & df['whatsapp'].duplicated(keep='first'))]
    
    return df


def save_data(df):
    """Save data to JSON and Excel"""
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    
    # Convert DataFrame to list of dicts
    data = df.to_dict('records')
    
    # Save JSON
    with open(output_dir / "affiliators_full.json", 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    # Save Excel with formatting
    save_to_excel(df)


def save_to_excel(df):
    """Save DataFrame to Excel with nice formatting"""
    if df.empty:
        return
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Affiliators"
    
    # Define column order
    column_order = [
        'index', 'username', 'display_name', 'shop_name', 'level', 'rating',
        'category', 'followers', 'bio', 'gmv', 'products_sold', 'avg_views',
        'gender_male', 'gender_female', 'age_group',
        'email', 'whatsapp', 'scraped_at'
    ]
    
    # Filter existing columns
    existing_columns = [col for col in column_order if col in df.columns]
    df_export = df[existing_columns].copy()
    
    # Write headers
    headers = {
        'index': 'No',
        'username': 'Username',
        'display_name': 'Display Name',
        'shop_name': 'Shop Name',
        'level': 'Level',
        'rating': 'Rating',
        'category': 'Category',
        'followers': 'Followers',
        'bio': 'Bio',
        'gmv': 'GMV',
        'products_sold': 'Products Sold',
        'avg_views': 'Avg Views',
        'gender_male': 'Male %',
        'gender_female': 'Female %',
        'age_group': 'Age Group',
        'email': 'Email',
        'whatsapp': 'WhatsApp',
        'scraped_at': 'Scraped At'
    }
    
    # Write header row
    for col_idx, col in enumerate(existing_columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = headers.get(col, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_idx, row_data in enumerate(df_export.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Highlight rows with contact info
            if col_idx in [existing_columns.index('email')+1 if 'email' in existing_columns else -1,
                          existing_columns.index('whatsapp')+1 if 'whatsapp' in existing_columns else -1]:
                if value and pd.notna(value):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Set column widths
    column_widths = {
        'index': 8,
        'username': 20,
        'display_name': 20,
        'shop_name': 20,
        'level': 8,
        'rating': 15,
        'category': 30,
        'followers': 15,
        'bio': 40,
        'gmv': 15,
        'products_sold': 15,
        'avg_views': 15,
        'gender_male': 10,
        'gender_female': 10,
        'age_group': 12,
        'email': 30,
        'whatsapp': 18,
        'scraped_at': 20
    }
    
    for col_idx, col in enumerate(existing_columns, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = column_widths.get(col, 15)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(existing_columns)):
        for cell in row:
            cell.border = thin_border
    
    # Save
    excel_path = output_dir / "affiliators_full.xlsx"
    wb.save(excel_path)
    
    # Also create contacts-only sheet
    if 'email' in df.columns or 'whatsapp' in df.columns:
        contacts_df = df[(df['email'].notna()) | (df['whatsapp'].notna())]
        if not contacts_df.empty:
            wb_contacts = Workbook()
            ws_contacts = wb_contacts.active
            ws_contacts.title = "With Contacts"
            
            # Write to contacts sheet (same formatting)
            for col_idx, col in enumerate(existing_columns, 1):
                cell = ws_contacts.cell(row=1, column=col_idx)
                cell.value = headers.get(col, col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row_idx, row_data in enumerate(contacts_df[existing_columns].values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_contacts.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = thin_border
            
            for col_idx, col in enumerate(existing_columns, 1):
                ws_contacts.column_dimensions[chr(64 + col_idx)].width = column_widths.get(col, 15)
            
            ws_contacts.freeze_panes = 'A2'
            
            contacts_path = output_dir / "affiliators_with_contacts.xlsx"
            wb_contacts.save(contacts_path)


def get_statistics(df):
    """Calculate statistics from DataFrame"""
    if df.empty:
        return {
            'total': 0,
            'with_email': 0,
            'with_whatsapp': 0,
            'with_contact': 0,
            'levels': {},
            'categories': {},
            'avg_followers': 0
        }
    
    total = len(df)
    with_email = df['email'].notna().sum() if 'email' in df.columns else 0
    with_whatsapp = df['whatsapp'].notna().sum() if 'whatsapp' in df.columns else 0
    with_contact = ((df['email'].notna()) | (df['whatsapp'].notna())).sum() if 'email' in df.columns and 'whatsapp' in df.columns else 0
    
    # Level distribution
    levels = {}
    if 'level' in df.columns:
        levels = df['level'].value_counts().to_dict()
    
    # Category distribution (top 10)
    categories = {}
    if 'category' in df.columns:
        df['main_category'] = df['category'].str.split(',').str[0]
        categories = df['main_category'].value_counts().head(10).to_dict()
    
    # Average followers
    avg_followers = 0
    if 'followers' in df.columns:
        def parse_number(value):
            if not value or pd.isna(value):
                return 0
            value = str(value).lower().replace(',', '.')
            match = re.search(r'([\d\.]+)\s*(rb|jt|k|m)?', value)
            if not match:
                return 0
            number = float(match.group(1))
            unit = match.group(2) if match.group(2) else ''
            if unit in ['rb', 'k']:
                return number * 1000
            elif unit in ['jt', 'm']:
                return number * 1000000
            return number
        
        df['followers_num'] = df['followers'].apply(parse_number)
        avg_followers = int(df['followers_num'].mean())
    
    return {
        'total': total,
        'with_email': with_email,
        'with_whatsapp': with_whatsapp,
        'with_contact': with_contact,
        'levels': levels,
        'categories': categories,
        'avg_followers': avg_followers
    }


@app.route('/')
def index():
    """Main dashboard page"""
    return render_template('index.html')


@app.route('/api/stats')
def api_stats():
    """Get current statistics"""
    df = load_data()
    stats = get_statistics(df)
    return jsonify(stats)


@app.route('/api/data')
def api_data():
    """Get all data"""
    df = load_data()
    
    # Convert to list of dicts
    data = df.to_dict('records')
    
    return jsonify(data)


@app.route('/api/scrape', methods=['POST'])
def api_scrape():
    """Start scraping"""
    global scraping_status
    
    if scraping_status['is_running']:
        return jsonify({'error': 'Scraping already in progress'}), 400
    
    data = request.json
    max_creators = data.get('max_creators', 10)
    
    # Start scraping in background thread
    def run_scraper():
        global scraping_status
        scraping_status['is_running'] = True
        scraping_status['progress'] = 0
        scraping_status['total'] = max_creators
        scraping_status['message'] = 'Starting scraper...'
        
        try:
            # Run scraper
            result = subprocess.run(
                ['python', 'scrape_full_data.py', '--max-creators', str(max_creators)],
                capture_output=True,
                text=True
            )
            
            if result.returncode == 0:
                # Load new data and remove duplicates
                df = load_data()
                df = remove_duplicates(df)
                save_data(df)
                
                scraping_status['message'] = f'Completed! Scraped {len(df)} creators'
            else:
                scraping_status['message'] = f'Error: {result.stderr[:200]}'
        
        except Exception as e:
            scraping_status['message'] = f'Error: {str(e)}'
        
        finally:
            scraping_status['is_running'] = False
            scraping_status['progress'] = scraping_status['total']
    
    thread = threading.Thread(target=run_scraper)
    thread.start()
    
    return jsonify({'message': 'Scraping started'})


@app.route('/api/status')
def api_status():
    """Get scraping status"""
    return jsonify(scraping_status)


@app.route('/api/remove-duplicates', methods=['POST'])
def api_remove_duplicates():
    """Remove duplicate entries"""
    df = load_data()
    
    if df.empty:
        return jsonify({'message': 'No data to process'}), 400
    
    original_count = len(df)
    df = remove_duplicates(df)
    new_count = len(df)
    
    save_data(df)
    
    removed = original_count - new_count
    
    return jsonify({
        'message': f'Removed {removed} duplicates',
        'original_count': original_count,
        'new_count': new_count
    })


@app.route('/api/export/excel')
def api_export_excel():
    """Export to Excel"""
    excel_path = Path("output/affiliators_full.xlsx")
    
    if not excel_path.exists():
        return jsonify({'error': 'No data to export'}), 404
    
    return send_file(
        excel_path,
        as_attachment=True,
        download_name=f'affiliators_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/api/export/contacts')
def api_export_contacts():
    """Export contacts only"""
    contacts_path = Path("output/affiliators_with_contacts.xlsx")
    
    if not contacts_path.exists():
        return jsonify({'error': 'No contacts data to export'}), 404
    
    return send_file(
        contacts_path,
        as_attachment=True,
        download_name=f'affiliators_contacts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


if __name__ == '__main__':
    import os
    
    # Create directories
    Path("output").mkdir(exist_ok=True)
    Path("templates").mkdir(exist_ok=True)
    
    # Get port from environment (for cloud platforms)
    port = int(os.environ.get('PORT', 5000))
    
    # Production mode
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    print("\n" + "="*80)
    print("🚀 TOKOPEDIA AFFILIATOR DASHBOARD")
    print("="*80)
    print(f"\n📊 Dashboard running on port: {port}")
    print(f"🔧 Debug mode: {debug}")
    print("\n⚠️  IMPORTANT:")
    print("   - Solve CAPTCHA manually when scraping")
    print("   - Dashboard will show progress")
    print("   - Data auto-saved to Excel\n")
    print("="*80 + "\n")
    
    app.run(debug=debug, host='0.0.0.0', port=port)
